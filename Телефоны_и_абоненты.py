# Библиотеки
import sys
import openpyxl
from PyQt6.QtWidgets import QApplication, QMainWindow, QTabWidget, QWidget, \
    QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout, QPushButton, \
    QFormLayout, QLabel, QLineEdit, QComboBox, QDateEdit, QMessageBox, QDialog, \
    QDoubleSpinBox, QTextEdit
from PyQt6.QtCore import Qt, QDate
import sqlite3
from openpyxl import Workbook

# Класс главного и диалоговых окон
class MainWindow(QMainWindow):
    # Функция "конструктор"
    def __init__(self):
        super().__init__()

        self.db = sqlite3.connect('phones_abonents.db')
        self.cursor = self.db.cursor()
        self.setGeometry(100,100,1000,500)
        self.create_tables()

        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        self.abonents_tab = QWidget()
        self.phones_tab = QWidget()
        self.plans_tab = QWidget()
        self.import_tab=QWidget()
        self.export_tab = QWidget()

        self.tab_widget.addTab(self.abonents_tab, 'Абоненты')
        self.tab_widget.addTab(self.phones_tab, 'Телефоны')
        self.tab_widget.addTab(self.plans_tab, 'Тарифы')
        self.tab_widget.addTab(self.import_tab, 'Импорт')
        self.tab_widget.addTab(self.export_tab, 'Экспорт')
        self.create_abonents_tab()
        self.create_phones_tab()
        self.create_plans_tab()
        self.create_import_tab()
        self.create_export_tab()

        self.update_plans_table()
        self.update_phones_table()
        self.update_abonents_table()

    # Создание таблицы если ее нет
    def create_tables(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS abonents (
                id_abonent INTEGER PRIMARY KEY,
                name VARCHAR,
                birth VARCHAR,
                entity VARCHAR
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS phones (
                id_phone INTEGER PRIMARY KEY,
                owner_id INTEGER,
                plan_id INTEGER,
                phone INTEGER,
                region VARCHAR,
                block VARCHAR,
                roaming VARCHAR,
                lastactive DATE,
                regdate DATE,
                FOREIGN KEY (owner_id) REFERENCES abonents (id_abonent),
                FOREIGN KEY (plan_id) REFERENCES plans (id_plan)
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS plans (
                id_plan INTEGER PRIMARY KEY,
                name VARCHAR,
                price INTEGER,
                traffic VARCHAR,
                calls VARCHAR,
                sms VARCHAR
            )
        ''')

    # Создание вкладки абонентов при запуске
    def create_abonents_tab(self):
        layout = QVBoxLayout()
        self.abonents_table = QTableWidget()
        self.abonents_table.setSortingEnabled(True)
        self.abonents_table.setColumnCount(4)
        self.abonents_table.setHorizontalHeaderLabels(['ID', 'Имя', 'Дата рождения', 'Тип субъекта'])
        layout.addWidget(self.abonents_table)

        create_button = QPushButton('Создать')
        create_button.clicked.connect(self.create_abonent)
        layout.addWidget(create_button)

        edit_button = QPushButton("Редактировать")
        edit_button.clicked.connect(self.edit_abonent)
        layout.addWidget(edit_button)

        delete_button = QPushButton('Удалить')
        delete_button.clicked.connect(self.delete_abonent)
        layout.addWidget(delete_button)

        self.abonents_tab.setLayout(layout)

        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Поиск")
        layout.addWidget(search_edit)
        search_edit.textChanged.connect(lambda: self.apply_filter(self.abonents_table, search_edit.text()))

    # Создание вкладки телефонов при запуске
    def create_phones_tab(self):
        layout = QVBoxLayout()
        self.phones_table = QTableWidget()
        self.phones_table.setSortingEnabled(True)
        self.phones_table.setColumnCount(9)
        self.phones_table.setHorizontalHeaderLabels(['ID', 'Владелец', 'Тариф', 'Номер', 'Регион', 'Блокировка', 'Роуминг', 'Последняя активность', 'Дата регистрации'])
        layout.addWidget(self.phones_table)

        create_button = QPushButton('Создать')
        create_button.clicked.connect(self.create_phone)
        layout.addWidget(create_button)

        edit_button = QPushButton("Редактировать")
        edit_button.clicked.connect(self.edit_phone)
        layout.addWidget(edit_button)

        delete_button = QPushButton('Удалить')
        delete_button.clicked.connect(self.delete_phone)
        layout.addWidget(delete_button)

        self.phones_tab.setLayout(layout)

        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Поиск")
        layout.addWidget(search_edit)
        search_edit.textChanged.connect(
            lambda: self.apply_filter(self.phones_table, search_edit.text()))

    # Создание вкладки тарифы при запуске
    def create_plans_tab(self):
        layout = QVBoxLayout()
        self.plans_table = QTableWidget()
        self.plans_table.setSortingEnabled(True)
        self.plans_table.setColumnCount(6)
        self.plans_table.setHorizontalHeaderLabels(['ID', 'Название', 'Цена', 'Трафик', 'Звонки', 'СМС'])
        layout.addWidget(self.plans_table)

        create_button = QPushButton('Создать')
        create_button.clicked.connect(self.create_plan)
        layout.addWidget(create_button)

        edit_button = QPushButton("Редактировать")
        edit_button.clicked.connect(self.edit_plan)
        layout.addWidget(edit_button)

        delete_button = QPushButton('Удалить')
        delete_button.clicked.connect(self.delete_plan)
        layout.addWidget(delete_button)

        self.plans_tab.setLayout(layout)

        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Поиск")
        layout.addWidget(search_edit)
        search_edit.textChanged.connect(
            lambda: self.apply_filter(self.plans_table, search_edit.text()))

    # Создание вкладки импорта в главном окне
    def create_import_tab(self):
        layout=QVBoxLayout()
        import_abonents_btn=QPushButton('Импортировать Excel таблицу "abonents.xlsx" в таблицу абонентов ')
        import_phones_btn=QPushButton('Импортировать Excel таблицу "phones.xlsx" в таблицу телефонов')
        import_plans_btn=QPushButton('Импортировать Excel таблицу "plans.xlsx" в таблицу тарифов')
        layout.addWidget(import_abonents_btn)
        layout.addWidget(import_phones_btn)
        layout.addWidget(import_plans_btn)
        import_abonents_btn.clicked.connect(self.import_abonents_to_excel)
        import_phones_btn.clicked.connect(self.import_phones_to_excel)
        import_plans_btn.clicked.connect(self.import_plans_to_excel)
        self.import_tab.setLayout(layout)

    # Создание вкладки экспорта в главном окне
    def create_export_tab(self):
        layout=QVBoxLayout()
        export_abonents_btn=QPushButton('Экспортировать таблицу абонентов в Excel')
        export_phones_btn=QPushButton('Экспортировать таблицу телефонов в Excel')
        export_plans_btn=QPushButton('Экспортировать таблицу тарифов в Excel')
        layout.addWidget(export_abonents_btn)
        layout.addWidget(export_phones_btn)
        layout.addWidget(export_plans_btn)
        export_abonents_btn.clicked.connect(self.export_abonents_to_excel)
        export_phones_btn.clicked.connect(self.export_phones_to_excel)
        export_plans_btn.clicked.connect(self.export_plans_to_excel)
        self.export_tab.setLayout(layout)

    # Создание абонента диалоговым окном
    def create_abonent(self):
        dialog = QDialog()
        dialog.setWindowTitle('Создать абонента')

        layout = QFormLayout()
        name_edit = QLineEdit()
        birth_edit = QDateEdit()
        entity_edit = QLineEdit()
        layout.addRow('Имя', name_edit)
        layout.addRow('Дата рождения', birth_edit)
        layout.addRow('Тип субъекта', entity_edit)
        create_button = QPushButton('Создать')
        create_button.clicked.connect(lambda: self.create_abonent_slot(name_edit.text(), birth_edit.date().toString('yyyy-MM-dd'), entity_edit.text(), dialog))
        layout.addWidget(create_button)

        dialog.setLayout(layout)
        dialog.exec()

    # Заполнение данных абонента в таблицу
    def create_abonent_slot(self, name, birth, entity, dialog):
        if not name or not entity:
            QMessageBox.warning(self, 'Ошибка', 'Заполните все поля')
            return

        try:
            self.cursor.execute('INSERT INTO abonents (name, birth, entity) VALUES (?, ?, ?)', (name, birth, entity))
            self.db.commit()
            self.update_abonents_table()
            QMessageBox.information(self, 'Успешно', 'Абонент создан')
            dialog.close()
        except sqlite3.Error as e:
            QMessageBox.critical(self, 'Ошибка', str(e))

    # Создание телефона диалоговым окном
    def create_phone(self):
        dialog = QDialog()
        dialog.setWindowTitle('Создать телефон')

        layout = QFormLayout()
        owner_id_combo = QComboBox()
        owner_id_combo.addItems([str(row[0]) for row in self.cursor.execute('SELECT id_abonent FROM abonents').fetchall()])
        plan_id_combo = QComboBox()
        plan_id_combo.addItems([str(row[0]) for row in self.cursor.execute('SELECT id_plan FROM plans').fetchall()])

        if not owner_id_combo and not plan_id_combo:
            QMessageBox.warning(self, 'Ошибка', 'Сначала зарегистрируйте абонента и создайте тариф')
            return
        if not owner_id_combo:
            QMessageBox.warning(self, 'Ошибка', 'Сначала зарегистрируйте абонента')
            return
        if not plan_id_combo:
            QMessageBox.warning(self, 'Ошибка', 'Сначала создайте тариф')
            return

        phone_edit = QLineEdit()
        region_edit = QLineEdit()
        block_edit = QLineEdit()
        roaming_edit = QLineEdit()
        lastactive_edit = QDateEdit()
        regdate_edit = QDateEdit()
        layout.addRow('Владелец', owner_id_combo)
        layout.addRow('Тариф', plan_id_combo)
        layout.addRow('Номер', phone_edit)
        layout.addRow('Регион', region_edit)
        layout.addRow('Блокировка', block_edit)
        layout.addRow('Роуминг', roaming_edit)
        layout.addRow('Последняя активность', lastactive_edit)
        layout.addRow('Дата регистрации', regdate_edit)

        create_button = QPushButton('Создать')
        create_button.clicked.connect(lambda: self.create_phone_slot(owner_id_combo.currentText(), plan_id_combo.currentText(), phone_edit.text(), region_edit.text(), block_edit.text(), roaming_edit.text(), lastactive_edit.date().toString('yyyy-MM-dd'), regdate_edit.date().toString('yyyy-MM-dd'), dialog))
        layout.addWidget(create_button)

        dialog.setLayout(layout)
        dialog.exec()

    # Заполнение данных телефона в таблицу
    def create_phone_slot(self, owner_id, plan_id, phone, region, block, roaming, lastactive, regdate, dialog):
        if not owner_id or not plan_id or not phone or not region or not block or not roaming or not  lastactive or not regdate:
            QMessageBox.warning(self, 'Ошибка', 'Заполните все поля')
            return
        try:
            phone=int(phone)
            if len(str(phone))!=11:
                QMessageBox.warning(self, 'Ошибка', 'Не корректная длина номера телефона')
                return
        except Exception:
            QMessageBox.warning(self, 'Ошибка', 'Не корректный формат телефона')
            return
        try:
            self.cursor.execute('INSERT INTO phones (owner_id, plan_id, phone, region, block, roaming, lastactive, regdate) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', (owner_id, plan_id, phone, region, block, roaming, lastactive, regdate))
            self.db.commit()
            self.update_phones_table()
            QMessageBox.information(self, 'Успешно', 'Телефон создан')
            dialog.close()
        except sqlite3.Error as e:
            QMessageBox.critical(self, 'Ошибка', str(e))

    # Создание тарифа диалоговым окном
    def create_plan(self):
        dialog = QDialog()
        dialog.setWindowTitle('Создать тариф')

        layout = QFormLayout()
        name_edit = QLineEdit()
        price_edit = QLineEdit()
        traffic_edit = QLineEdit()
        calls_edit = QLineEdit()
        sms_edit = QLineEdit()
        layout.addRow('Название', name_edit)
        layout.addRow('Цена', price_edit)
        layout.addRow('Трафик', traffic_edit)
        layout.addRow('Звонки', calls_edit)
        layout.addRow('СМС', sms_edit)

        create_button = QPushButton('Создать')
        create_button.clicked.connect(lambda: self.create_plan_slot(name_edit.text(), price_edit.text(), traffic_edit.text(), calls_edit.text(), sms_edit.text(), dialog))
        layout.addWidget(create_button)

        dialog.setLayout(layout)
        dialog.exec()

    # Заполнение данных тарифа в таблицу
    def create_plan_slot(self, name, price, traffic, calls, sms, dialog):
        if not name or not price or not traffic or not calls or not sms:
            QMessageBox.warning(self, 'Ошибка', 'Заполните все поля')
            return
        try:
            price=int(price)
        except Exception:
            QMessageBox.warning(self, 'Ошибка', 'Неверный формат цены')
            return
        try:
            self.cursor.execute('INSERT INTO plans (name, price, traffic, calls, sms) VALUES (?, ?, ?, ?, ?)', (name, price, traffic, calls, sms))
            self.db.commit()
            self.update_plans_table()
            QMessageBox.information(self, 'Успешно', 'Тариф создан')
            dialog.close()
        except sqlite3.Error as e:
            QMessageBox.critical(self, 'Ошибка', str(e))

    # Редактирование абонента
    def edit_abonent(self):
        try:
            selected_row = self.abonents_table.currentRow()
            if selected_row == -1:
                QMessageBox.warning(self, 'Предупреждение', 'Выберите абонента')
                return

            abonent_data = self.cursor.execute("SELECT * FROM abonents WHERE id_abonent=?",
                                               (self.abonents_table.item(selected_row,0).text(),)).fetchone()
            dialog = QDialog()
            dialog.setWindowTitle("Редактировать абонента")

            layout = QFormLayout()
            name_edit = QLineEdit(abonent_data[1])
            birth_edit = QLineEdit(abonent_data[2])
            entity_edit = QLineEdit(abonent_data[3])
            layout.addRow("Имя", name_edit)
            layout.addRow("Дата рождения", birth_edit)
            layout.addRow("Тип субъекта", entity_edit)

            save_button = QPushButton("Сохранить")
            save_button.clicked.connect(
                lambda: self.save_abonent_edit(abonent_data[0], name_edit.text(),birth_edit.text(),
                                               entity_edit.text(), dialog))
            layout.addWidget(save_button)

            dialog.setLayout(layout)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка редактирования", str(e))

    # Обновление данных абонента
    def save_abonent_edit(self, abonent_id, name, birth, entity, dialog):
        try:
            if not name or not entity:
                QMessageBox.warning(self, 'Ошибка', 'Заполните все поля')
                return
            self.cursor.execute(
                "UPDATE abonents SET name=?, birth=?, entity=? WHERE id_abonent=?",
                (name, birth, entity, abonent_id))
            self.db.commit()


            self.update_abonents_table()

            QMessageBox.information(self, "Успешно", "Абонент обновлен")
            dialog.close()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка обновления", str(e))

    # Редактирование телефона
    def edit_phone(self):
        try:
            selected_row = self.phones_table.currentRow()
            if selected_row == -1:
                QMessageBox.warning(self, 'Предупреждение', 'Выберите телефон')
                return

            phone_data = self.cursor.execute("SELECT * FROM phones WHERE id_phone=?",
                                             (self.phones_table.item(selected_row,
                                                                     0).text(),)).fetchone()

            dialog = QDialog()
            dialog.setWindowTitle("Редактировать телефон")

            layout = QFormLayout()

            owner_id_combo = QComboBox()
            owner_id_combo.addItems([str(row[0]) for row in self.cursor.execute(
                'SELECT id_abonent FROM abonents').fetchall()])
            owner_id_combo.setCurrentIndex(owner_id_combo.findText(str(phone_data[1])))
            plan_id_combo = QComboBox()
            plan_id_combo.addItems([str(row[0]) for row in self.cursor.execute(
                'SELECT id_plan FROM plans').fetchall()])
            plan_id_combo.setCurrentIndex(plan_id_combo.findText(str(phone_data[2])))

            phone_edit = QLineEdit(str(phone_data[3]))
            region_edit = QLineEdit(str(phone_data[4]))
            block_edit = QLineEdit(str(phone_data[5]))
            roaming_edit = QLineEdit(str(phone_data[6]))
            lastactive_edit = QDateEdit()
            lastactive_edit.setDate(QDate.fromString(phone_data[7], "yyyy-MM-dd"))
            regdate_edit = QDateEdit()
            regdate_edit.setDate(QDate.fromString(phone_data[8], "yyyy-MM-dd"))

            layout.addRow("Владелец", owner_id_combo)
            layout.addRow("Тариф", plan_id_combo)
            layout.addRow("Номер", phone_edit)
            layout.addRow("Регион", region_edit)
            layout.addRow("Блокировка", block_edit)
            layout.addRow("Роуминг", roaming_edit)
            layout.addRow("Последняя активность", lastactive_edit)
            layout.addRow("Дата регистрации", regdate_edit)

            save_button = QPushButton("Сохранить")
            save_button.clicked.connect(lambda: self.save_phone_edit(phone_data[0],
                                                                     owner_id_combo.currentText(),
                                                                     plan_id_combo.currentText(),
                                                                     phone_edit.text(),
                                                                     region_edit.text(),
                                                                     block_edit.text(),
                                                                     roaming_edit.text(),
                                                                     lastactive_edit.date().toString(
                                                                         "yyyy-MM-dd"),
                                                                     regdate_edit.date().toString(
                                                                         "yyyy-MM-dd"),
                                                                     dialog))
            layout.addWidget(save_button)

            dialog.setLayout(layout)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    # Обновление данных телефона
    def save_phone_edit(self, phone_id, owner_id, plan_id, phone, region, block,roaming, lastactive, regdate, dialog):
        try:
            if not owner_id or not plan_id or not phone or not region or not block or not roaming or not lastactive or not regdate:
                QMessageBox.warning(self, 'Ошибка', 'Заполните все поля')
                return
            try:
                phone = int(phone)
                if len(str(phone)) != 11:
                    QMessageBox.warning(self, 'Ошибка',
                                        'Не корректная длина номера телефона')
                    return
            except Exception:
                QMessageBox.warning(self, 'Ошибка',
                                    'Не корректный формат телефона')
                return
            self.cursor.execute(
                "UPDATE phones SET owner_id=?, plan_id=?, phone=?, region=?, block=?, roaming=?, lastactive=?, regdate=? WHERE id_phone=?",
                (owner_id, plan_id, phone, region, block, roaming, lastactive,
                 regdate, phone_id))
            self.db.commit()

            self.update_phones_table()

            dialog.close()

            QMessageBox.information(self, "Успешно", "Телефон обновлен")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    # Редактирование тарифа
    def edit_plan(self):
        try:
            selected_row = self.plans_table.currentRow()
            if selected_row == -1:
                QMessageBox.warning(self, 'Предупреждение', 'Выберите тариф')
                return

            plan_data = self.cursor.execute("SELECT * FROM plans WHERE id_plan=?",
                                            (self.plans_table.item(selected_row,
                                                                   0).text(),)).fetchone()

            dialog = QDialog()
            dialog.setWindowTitle("Редактировать тариф")

            layout = QFormLayout()

            name_edit = QLineEdit(plan_data[1])
            price_edit = QDoubleSpinBox()
            price_edit.setRange(0.1, 1000000.0)
            price_edit.setValue(plan_data[2])
            traffic_edit = QLineEdit(plan_data[3])
            calls_edit = QLineEdit(plan_data[4])
            sms_edit = QLineEdit(plan_data[5])

            layout.addRow("Название", name_edit)
            layout.addRow("Цена", price_edit)
            layout.addRow("Трафик", traffic_edit)
            layout.addRow("Звонки", calls_edit)
            layout.addRow("Сообщения", sms_edit)

            save_button = QPushButton("Сохранить")
            save_button.clicked.connect(
                lambda: self.save_plan_edit(plan_data[0], name_edit.text(),
                                            price_edit.value(),
                                            traffic_edit.text(),
                                            calls_edit.text(), sms_edit.text(),
                                            dialog))
            layout.addWidget(save_button)

            dialog.setLayout(layout)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка редактирования", str(e))

    # Обновление данных тарифа
    def save_plan_edit(self, plan_id, name, price, traffic, calls,sms, dialog):
        try:
            if not name or not price or not traffic or not calls or not sms:
                QMessageBox.warning(self, 'Ошибка', 'Заполните все поля')
                return
            self.cursor.execute(
                "UPDATE plans SET name=?, price=?, traffic=?, calls=?, sms=? WHERE id_plan=?",
                (name, price, traffic, calls,sms, plan_id))
            self.db.commit()

            self.update_plans_table()

            dialog.close()
            QMessageBox.information(self, "Успешно", "Тариф обновлен")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    # Импорт абонентов из Excel
    def import_abonents_to_excel(self):
        try:
            reply = QMessageBox.question(self, 'Подтверждение',
                                         'Начать импорт таблицы абонентов?',
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                return
            reply = QMessageBox.question(self, 'Параметры импорта',
                                         'Заменить текущую таблицу?',
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                         QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                self.cursor.execute('DELETE FROM abonents')
            wb = openpyxl.load_workbook("abonents.xlsx")
            ws1 = wb.active

            abonents_data = []
            for row in ws1.rows:
                abonents_data.append([cell.value for cell in row])

            for row in abonents_data:
                self.cursor.execute('''
                    INSERT INTO abonents (name, birth, entity)
                    VALUES (?, ?, ?)
                ''', row[1:])
                self.db.commit()
                self.update_abonents_table()
            QMessageBox.information(self, 'Успешно',
                                    'Таблица абонентов импортирована из Excel')
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', str(e))

    # Экспорт абонентов в Excel
    def export_abonents_to_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            rows = self.cursor.execute('SELECT * FROM abonents').fetchall()
            for i, row in enumerate(rows):
                for j, item in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1).value = item
            wb.save('abonents.xlsx')
            QMessageBox.information(self, 'Успешно',
                                    'Таблица абонентов экспортирована в Excel')
        except Exception:
            QMessageBox.critical(self, 'Внимание',
                                    'Сначала закройте Excel таблицу абонентов')

    # Импорт телефонов из Excel
    def import_phones_to_excel(self):
        try:
            reply = QMessageBox.question(self, 'Подтверждение',
                                         'Начать импорт таблицы телефонов?',
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                         QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                return
            reply = QMessageBox.question(self, 'Параметры импорта',
                                         'Заменить текущую таблицу?',
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                         QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                self.cursor.execute('DELETE FROM phones')
            wb = openpyxl.load_workbook("phones.xlsx")
            ws1 = wb.active

            phones_data = []
            for row in ws1.rows:
                phones_data.append([cell.value for cell in row])

            for row in phones_data:
                self.cursor.execute('''
                    INSERT INTO phones (owner_id, plan_id, phone, region, block, roaming, lastactive, regdate)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', row[1:])
                self.db.commit()
                self.update_phones_table()
            QMessageBox.information(self, 'Успешно',
                                    'Таблица номеров импортирована из Excel')
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', str(e))

    # Экспорт телефонов в Excel
    def export_phones_to_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            rows = self.cursor.execute('SELECT * FROM phones').fetchall()
            for i, row in enumerate(rows):
                for j, item in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1).value = item
            wb.save('phones.xlsx')
            QMessageBox.information(self, 'Успешно',
                                    'Таблица телефонов экспортирована в Excel')
        except Exception:
            QMessageBox.critical(self, 'Внимание',
                                    'Сначала закройте Excel таблицу телефонов')

    # Импорт тарифов из Excel
    def import_plans_to_excel(self):
        try:
            reply = QMessageBox.question(self, 'Подтверждение',
                                         'Начать импорт таблицы тарифов?',
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                         QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                return
            reply = QMessageBox.question(self, 'Параметры импорта',
                                         'Заменить текущую таблицу?',
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                         QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                self.cursor.execute('DELETE FROM plans')

            wb = openpyxl.load_workbook("plans.xlsx")
            ws1 = wb.active

            plans_data = []
            for row in ws1.rows:
                plans_data.append([cell.value for cell in row])

            for row in plans_data:
                self.cursor.execute('''
                    INSERT INTO plans (name, price, traffic, calls, sms)
                    VALUES (?, ?, ?, ?, ?)
                ''', row[1:])
                self.db.commit()
                self.update_plans_table()
            QMessageBox.information(self, 'Успешно',
                                    'Таблица тарифов импортирована из Excel')
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', str(e))

    # Экспорт тарифов в Excel
    def export_plans_to_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            rows = self.cursor.execute('SELECT * FROM plans').fetchall()
            for i, row in enumerate(rows):
                for j, item in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1).value = item
            wb.save('plans.xlsx')
            QMessageBox.information(self, 'Успешно',
                                    'Таблица тарифов экспортирована в Excel')
        except Exception:
            QMessageBox.critical(self, 'Внимание',
                                    'Сначала закройте Excel тарифов телефонов')

    # Фильтр таблицы
    def apply_filter(self, table_widget, search_text):
        for row in range(table_widget.rowCount()):
            visible = False
            for column in range(table_widget.columnCount()):
                item = table_widget.item(row, column)
                if item and search_text.lower() in item.text().lower():
                    visible = True
                    break
            table_widget.setRowHidden(row, not visible)

    # Удаление записи об абоненте
    def delete_abonent(self):
        try:
            row = self.abonents_table.currentRow()
            if row >= 0:
                id_abonent = self.abonents_table.item(row, 0).text()
                if self.check_dependencies('abonents', id_abonent):
                    reply = QMessageBox.question(self, 'Удаление записи',
                                                 'Удалить запись и все связанные с ней записи в таблице Телефоны?',
                                                 QMessageBox.StandardButton.Yes,
                                                 QMessageBox.StandardButton.No)
                    if reply == QMessageBox.StandardButton.Yes:
                        self.cursor.execute('DELETE FROM abonents WHERE id_abonent = ?',
                                            (id_abonent,))
                        self.cursor.execute(
                            'DELETE FROM phones WHERE owner_id = ?',
                            (id_abonent,))
                        self.db.commit()
                    else:
                        reply = QMessageBox.question(self, 'Удаление записи',
                                                     'Удалить запись и маркировать связанные записи в таблице Телефоны?',
                                                     QMessageBox.StandardButton.Yes,
                                                     QMessageBox.StandardButton.No)
                        if reply == QMessageBox.StandardButton.Yes:
                            self.cursor.execute(
                                "UPDATE phones SET owner_id = '*' WHERE owner_id = ?",
                                (id_abonent,))
                            self.cursor.execute(
                                'DELETE FROM abonents WHERE id_abonent = ?',
                                (id_abonent,))
                            self.db.commit()
                        else:
                            return
                else:
                    reply = QMessageBox.question(self, 'Удаление записи',
                                                 'Вы действительно хотите удалить запись?',
                                                 QMessageBox.StandardButton.Yes,
                                                 QMessageBox.StandardButton.No)
                    if reply == QMessageBox.StandardButton.Yes:
                        self.cursor.execute(
                            'DELETE FROM abonents WHERE id_abonent = ?',
                            (id_abonent,))
                        self.db.commit()
            else:
                QMessageBox.warning(self, 'Ошибка',
                                    'Выберите запись для удаления.')
            QMessageBox.information(self, 'Успешно', 'Запись удалена')
            self.update_abonents_table()
            self.update_phones_table()
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка',str(e))

    # Удаление записи о телефоне
    def delete_phone(self):
        row = self.phones_table.currentRow()
        if row != -1:
            id_phone = self.phones_table.item(row, 0).text()
            confirm_dialog = QMessageBox()
            confirm_dialog.setWindowTitle('Предупреждение')
            confirm_dialog.setText('Вы действительно хотите удалить запись?')
            confirm_dialog.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            confirm_dialog.setDefaultButton(QMessageBox.StandardButton.No)
            if confirm_dialog.exec() == QMessageBox.StandardButton.Yes:
                self.cursor.execute('DELETE FROM phones WHERE id_phone = ?',
                                    (id_phone,))
                self.db.commit()
                self.update_phones_table()
        else:
            QMessageBox.warning(self, 'Предупреждение', 'Выберите телефон')
        QMessageBox.information(self, 'Успешно', 'Запись удалена')

    # Удаление записи о тарифе
    def delete_plan(self):
        try:
            row = self.plans_table.currentRow()
            if row >= 0:
                id_plan = self.plans_table.item(row, 0).text()
                if self.check_dependencies('plans', id_plan):
                    reply = QMessageBox.question(self, 'Удаление записи',
                                                 'Удалить запись и все связанные с ней записи в таблице Телефоны?',
                                                 QMessageBox.StandardButton.Yes,
                                                 QMessageBox.StandardButton.No)
                    if reply == QMessageBox.StandardButton.Yes:
                        self.cursor.execute(
                            'DELETE FROM plans WHERE id_plan = ?',
                            (id_plan,))
                        self.cursor.execute(
                            'DELETE FROM phones WHERE plan_id = ?',
                            (id_plan,))
                        self.db.commit()
                    else:
                        reply = QMessageBox.question(self, 'Удаление записи',
                                                     'Удалить запись и маркировать связанные записи в таблице Телефоны?',
                                                     QMessageBox.StandardButton.Yes,
                                                     QMessageBox.StandardButton.No)
                        if reply == QMessageBox.StandardButton.Yes:
                            self.cursor.execute(
                                "UPDATE phones SET plan_id = '*' WHERE plan_id = ?",
                                (id_plan,))
                            self.cursor.execute(
                                'DELETE FROM plans WHERE id = ?',
                                (id_plan,))
                            self.db.commit()
                        else:
                            return
                else:
                    reply = QMessageBox.question(self, 'Удаление записи',
                                                 'Вы действительно хотите удалить запись?',
                                                 QMessageBox.StandardButton.Yes,
                                                 QMessageBox.StandardButton.No)
                    if reply == QMessageBox.StandardButton.Yes:
                        self.cursor.execute('DELETE FROM plans WHERE id_plan = ?',
                                            (id_plan,))
                        self.db.commit()
            else:
                QMessageBox.warning(self, 'Ошибка',
                                    'Выберите запись для удаления.')
            QMessageBox.information(self, 'Успешно', 'Запись удалена')
            self.update_phones_table()
            self.update_plans_table()
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', str(e))

    # Проверка зависимостей перед удалением
    def check_dependencies(self, table_name, id):
        if table_name == 'abonents':
            query = "SELECT * FROM phones WHERE owner_id = ?"
        elif table_name == 'plans':
            query = "SELECT * FROM phones WHERE plan_id = ?"
        else:
            return False

        self.cursor.execute(query, (id,))
        if self.cursor.fetchone():
            return True
        else:
            return False

    # Обновление данных в таблице абонентов
    def update_abonents_table(self):
        self.abonents_table.clearContents()
        self.abonents_table.setRowCount(0)
        rows = self.cursor.execute('SELECT * FROM abonents').fetchall()
        for row in rows:
            self.abonents_table.setRowCount(self.abonents_table.rowCount() + 1)
            for i, item in enumerate(row):
                self.abonents_table.setItem(self.abonents_table.rowCount() - 1, i, QTableWidgetItem(str(item)))

    # Обновление данных в таблице телефонов
    def update_phones_table(self):
        self.phones_table.clearContents()
        self.phones_table.setRowCount(0)
        rows = self.cursor.execute('SELECT * FROM phones').fetchall()
        for row in rows:
            self.phones_table.setRowCount(self.phones_table.rowCount() + 1)
            for i, item in enumerate(row):
                self.phones_table.setItem(self.phones_table.rowCount() - 1, i, QTableWidgetItem(str(item)))

    # Обновление данных в таблице тарифов
    def update_plans_table(self):
        self.plans_table.clearContents()
        self.plans_table.setRowCount(0)
        rows = self.cursor.execute('SELECT * FROM plans').fetchall()
        for row in rows:
            self.plans_table.setRowCount(self.plans_table.rowCount() + 1)
            for i, item in enumerate(row):
                self.plans_table.setItem(self.plans_table.rowCount() - 1, i, QTableWidgetItem(str(item)))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
